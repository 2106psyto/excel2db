from openpyxl import load_workbook, utils
from .models import WorksheetConfig
from typing import List, Any


def _get_merged_headers(ws, number_of_rows):
    merged_headers = []
    max_col = ws.max_column

    for col_idx in range(1, max_col + 1):
        column_parts = []
        for row_num in range(1, number_of_rows + 1):
            cell_value = ws.cell(row=row_num, column=col_idx).value

            # NOTE:
            # openpyxl returns None as cell value except for the top-left one when
            # cells merged across multiple rows/columns.
            if cell_value is not None:
                column_parts.append(str(cell_value))

        if column_parts:
            merged_headers.append("_".join(column_parts))

    return merged_headers

def read_worksheet_data(excel_path: str, ws_config: WorksheetConfig) -> List[List[Any]]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[ws_config.name]
    data = []
    start_row, start_col = utils.coordinate_to_tuple(ws_config.start_cell)
    for row in ws.iter_rows(
        min_row=start_row,
        min_col=start_col,
        values_only=True
    ):
        data.append(list(row))
    header = _get_merged_headers(ws, ws_config.header_rows)
    return data, header